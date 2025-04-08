from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
import locale
import json
from parser_csv import parser  # Ваш парсер CSV


class Excel():
    def __init__(self, path, path_to_xlsx):
        self.csv = parser.parse(path)
        with open("settings.json", "r", encoding='utf-8') as setts:
            self.setts_json = json.load(setts)
        try:
            self.path_to_xlsx = path_to_xlsx
            self.wb = Workbook()
            self.wb = load_workbook(self.path_to_xlsx)

        except:
            self.wb = Workbook()
            self.Sheet1 = self.wb['Sheet']
            self.wb.remove(self.Sheet1)

    def create_list(self):
        locale.setlocale(category=locale.LC_ALL, locale="Russian")
        self.a = datetime.today()
        sheet = self.wb.create_sheet(title=self.a.strftime('%B')+'_предметы')

        # Устанавливаем ширину первой колонки и заголовок
        sheet.column_dimensions['A'].width = len(
            max(self.csv["Предметы"].keys()))
        sheet['A1'] = "Дисциплина"
        sheet['A1'].alignment = Alignment(horizontal='center')

        j = 2
        for i in list(self.csv["Предметы"].keys()):
            sheet[f'A{j}'] = i
            sheet[f'A{j}'].alignment = Alignment(
                horizontal='center', vertical='center')
            j += 1

        list_subj = list(self.csv["Предметы"].keys())
        list_quest = list(self.csv["Предметы"][list_subj[0]].keys())
        last_col = 2          # +2, т.к. 1-я колонка уже занята дисциплинами
        # Определяем последнюю колонку для записи среднего
        for quest in list(self.setts_json["Questions_for_subject"].keys()):
            if len(self.setts_json["Questions_for_subject"][quest]) > 1:
                last_col += 1

        # Заполняем заголовки вопросов и данных
        # Начинаем со 2-й колонки
        for col_idx, question in enumerate(list_quest, start=2):
            if (len(self.setts_json['Questions_for_subject'][question]) > 1):
                sheet.cell(row=1, column=col_idx, value=question).alignment = Alignment(
                    horizontal='center')

        # Заполняем данные и считаем среднее, учитывая только числовые ответы
        # Начинаем со 2-й строки
        for row_idx, subject in enumerate(list_subj, start=2):
            values = []
            for col_idx, question in enumerate(list_quest, start=2):
                if (len(self.setts_json['Questions_for_subject'][question]) > 1):
                    scores = self.csv["Предметы"][subject][question]

                    # Фильтруем только числовые ответы
                    numeric_scores = [
                        ans for ans in scores if isinstance(ans, (int, float))]

                    # Если есть числовые ответы, считаем среднее
                    avg_score = sum(numeric_scores) / \
                        len(numeric_scores) if numeric_scores else 0
                    sheet.cell(row=row_idx, column=col_idx, value=avg_score).alignment = Alignment(
                        horizontal='center')
                    values.append(avg_score)

            # Записываем среднее значение по строке в последнюю колонку
            if values:
                sheet.cell(row=row_idx, column=last_col, value=sum(
                    values) / len(values)).alignment = Alignment(horizontal='center')

        # Добавляем заголовок для колонки со средним
        sheet.cell(row=1, column=last_col, value="Итог").alignment = Alignment(
            horizontal='center')

        avg_row = len(list_subj) + 2

        # Среднее значение для столбца "Итог"
        total_values = [
            sheet.cell(row=row_idx, column=last_col).value
            for row_idx in range(2, len(list_subj) + 2)
            if sheet.cell(row=row_idx, column=last_col).value is not None]

        total_avg = sum(total_values) / \
            len(total_values) if total_values else 0
        sheet.cell(row=avg_row, column=last_col,
                   value=total_avg).alignment = Alignment(horizontal='center')

        self.wb.save(self.path_to_xlsx)

    def create_teachers_list(self):
        locale.setlocale(category=locale.LC_ALL, locale="Russian")
        self.a = datetime.today()
        sheet = self.wb.create_sheet(
            title=self.a.strftime('%B')+"_преподаватели")

        sheet.column_dimensions['A'].width = len(
            max(self.csv["Преподаватели"].keys(), key=len))
        sheet['A1'] = "Преподаватель"
        sheet['A1'].alignment = Alignment(horizontal='center')

        j = 2
        for i in list(self.csv["Преподаватели"].keys()):
            sheet[f'A{j}'] = i
            sheet[f'A{j}'].alignment = Alignment(
                horizontal='center', vertical='center')
            j += 1

        list_teachers = list(self.csv["Преподаватели"].keys())
        list_quest = list(self.setts_json["Questions_for_teachers"].keys())

        last_col = 2
        for quest in list(self.setts_json["Questions_for_teachers"].keys()):
            if len(self.setts_json["Questions_for_teachers"][quest]) > 1:
                last_col += 1

        for col_idx, question in enumerate(list_quest, start=2):
            if (len(self.setts_json['Questions_for_teachers'][question]) > 1):
                sheet.cell(row=1, column=col_idx, value=question).alignment = Alignment(
                    horizontal='center')

        for row_idx, teacher in enumerate(list_teachers, start=2):
            values = []
            for col_idx, question in enumerate(list_quest, start=2):
                if (len(self.setts_json['Questions_for_teachers'][question]) > 1):
                    scores = self.csv["Преподаватели"][teacher][question]
                    numeric_scores = [
                        ans for ans in scores if isinstance(ans, (int, float))]

                    # Если есть числовые ответы, считаем среднее
                    avg_score = sum(numeric_scores) / \
                        len(numeric_scores) if numeric_scores else 0
                    sheet.cell(row=row_idx, column=col_idx, value=avg_score).alignment = Alignment(
                        horizontal='center')
                    values.append(avg_score)
            if values:
                sheet.cell(row=row_idx, column=last_col, value=sum(
                    values) / len(values)).alignment = Alignment(horizontal='center')

        sheet.cell(row=1, column=last_col, value="Итог").alignment = Alignment(
            horizontal='center')

        avg_row = len(list_teachers) + 2
        total_values = [sheet.cell(row=row_idx, column=last_col).value for row_idx in range(
            2, len(list_teachers) + 2) if sheet.cell(row=row_idx, column=last_col).value is not None]
        total_avg = sum(total_values) / \
            len(total_values) if total_values else 0
        sheet.cell(row=avg_row, column=last_col,
                   value=total_avg).alignment = Alignment(horizontal='center')

        self.wb.save(self.path_to_xlsx)


# nw = Excel("1_copy.csv", "1.xlsx")
# nw.create_list()
# nw.create_teachers_list()
